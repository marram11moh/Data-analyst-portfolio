

import pandas as pd
from openpyxl import load_workbook

SRC = "/mnt/user-data/uploads/eastern_province15.xlsx"
OUT = "/home/claude/clean_data"

wb = load_workbook(SRC, read_only=True)

GOVERNORATES = [
    "Dammam", "Al-Ahsa", "Hafar Al-Batin", "Jubail", "Qatif", "Khobar",
    "Khafji", "Ras Tannourah", "Abqaiq", "Nariya", "Qaryat al-Ulya", "Udayd",
]


def sheet_rows(sheet_name, max_row=60):
    """Return all non-empty rows of a sheet as a list of tuples."""
    ws = wb[sheet_name]
    return [row for row in ws.iter_rows(max_row=max_row, values_only=True)
            if any(c is not None for c in row)]


def rows_by_governorate(sheet_name, n_values):
    """
    Find the block of rows keyed by governorate name (col 0),
    stopping before the 'Total' row. Returns list of (governorate, *values).
    """
    ws = wb[sheet_name]
    out = []
    for row in ws.iter_rows(values_only=True):
        first = row[0]
        if isinstance(first, str) and first.strip() in GOVERNORATES:
            out.append((first.strip(), *row[1:1 + n_values]))
    return out


# ------------------------------------------------------------------
# 1) governorates (dimension table)

governorates_df = pd.DataFrame({
    "governorate_id": range(1, len(GOVERNORATES) + 1),
    "governorate_name": GOVERNORATES,
})
governorates_df.to_csv(f"{OUT}/governorates.csv", index=False)

gov_id = dict(zip(governorates_df.governorate_name, governorates_df.governorate_id))


def with_gov_id(df):
    df.insert(0, "governorate_id", df["governorate_name"].map(gov_id))
    return df.drop(columns="governorate_name")


# ------------------------------------------------------------------
# 2) population  (Sheet 3)

cols = ["governorate_name", "num_houses",
        "saudi_male", "saudi_female", "saudi_total",
        "non_saudi_male", "non_saudi_female", "non_saudi_total",
        "total_male", "total_female", "total_population"]
data = rows_by_governorate("3", n_values=10)
population_df = with_gov_id(pd.DataFrame(data, columns=cols))
population_df.to_csv(f"{OUT}/population.csv", index=False)

# ------------------------------------------------------------------
# 3) hospitals  (Sheet 13)

cols = ["governorate_name", "population",
        "govt_hospitals", "govt_beds", "govt_physicians",
        "private_hospitals", "private_beds", "private_physicians",
        "total_beds", "beds_per_1000", "total_physicians", "physicians_per_1000"]
data = rows_by_governorate("13", n_values=11)
hospitals_df = with_gov_id(pd.DataFrame(data, columns=cols))
hospitals_df.to_csv(f"{OUT}/hospitals.csv", index=False)

# ------------------------------------------------------------------
# 4) healthcare_centers  (Sheet 14)

cols = ["governorate_name", "population",
        "govt_centers", "govt_physicians",
        "private_dispensaries", "private_physicians",
        "red_crescent_centers", "total_physicians", "physicians_per_10000"]
data = rows_by_governorate("14", n_values=8)
healthcare_df = with_gov_id(pd.DataFrame(data, columns=cols))
healthcare_df.to_csv(f"{OUT}/healthcare_centers.csv", index=False)

# ------------------------------------------------------------------
# 5) higher_education  (Sheet 11)

cols = ["governorate_name",
        "state_universities", "state_university_branches",
        "state_students_male", "state_students_female",
        "state_staff_male", "state_staff_female",
        "private_universities",
        "private_students_male", "private_students_female",
        "private_staff_male", "private_staff_female"]
data = rows_by_governorate("11", n_values=11)
higher_ed_df = with_gov_id(pd.DataFrame(data, columns=cols))
higher_ed_df.to_csv(f"{OUT}/higher_education.csv", index=False)

# ------------------------------------------------------------------
# 6) education  (Sheets 4,5,6,7 -> melted into one tidy long table)
#    gender x sector x level (primary/intermediate/secondary)

LEVELS = ["Primary", "Intermediate", "Secondary"]
METRICS = ["schools", "classes", "students", "teachers",
           "avg_class_size", "avg_students_per_teacher"]

edu_sheets = {
    "4": ("Boys", "Government"),
    "5": ("Boys", "Private"),
    "6": ("Girls", "Government"),
    "7": ("Girls", "Private"),
}

edu_rows = []
for sheet_name, (gender, sector) in edu_sheets.items():
    data = rows_by_governorate(sheet_name, n_values=18)  # 3 levels x 6 metrics
    for row in data:
        governorate = row[0]
        values = row[1:]
        for level_idx, level in enumerate(LEVELS):
            level_vals = values[level_idx * 6:(level_idx + 1) * 6]
            edu_rows.append((governorate, gender, sector, level, *level_vals))

education_df = pd.DataFrame(
    edu_rows,
    columns=["governorate_name", "gender", "sector", "education_level", *METRICS],
)
education_df = with_gov_id(education_df)
education_df.to_csv(f"{OUT}/education.csv", index=False)

# ------------------------------------------------------------------
# 7) service_facilities_by_governorate  (Sheet 26)

INSTITUTIONS = [
    "real_estate_dev_fund", "industrial_dev_fund", "scsb_bank",
    "ministry_of_transport", "ministry_of_water", "ministry_of_finance",
    "ministry_of_planning", "ministry_of_justice", "retirement", "insurance",
]
cols = ["governorate_name"] + [f"{inst}_{kind}" for inst in INSTITUTIONS for kind in ("branch", "office")]
data = rows_by_governorate("26", n_values=20)
service_fac_df = with_gov_id(pd.DataFrame(data, columns=cols))
service_fac_df.to_csv(f"{OUT}/service_facilities_by_governorate.csv", index=False)

# ------------------------------------------------------------------
# 8) economic_establishments  (Sheet 1, region-level, by economic activity)

ws = wb["1"]
rows = [r[:5] for r in ws.iter_rows(min_row=10, max_row=27, values_only=True) if r[0]]
econ_estab_df = pd.DataFrame(
    rows, columns=["economic_activity", "private", "public", "non_profit", "total"]
)
econ_estab_df.to_csv(f"{OUT}/economic_establishments.csv", index=False)

# ------------------------------------------------------------------
# 9) economic_indicators  (Sheet 2, region-level, by economic activity, thousand SAR)

ws = wb["2"]
rows = [r[:4] for r in ws.iter_rows(min_row=9, max_row=26, values_only=True) if r[0]]
econ_ind_df = pd.DataFrame(
    rows, columns=["economic_activity", "revenue_thousand_sar",
                   "expenditure_thousand_sar", "operating_surplus_thousand_sar"]
)
econ_ind_df.to_csv(f"{OUT}/economic_indicators.csv", index=False)

# ------------------------------------------------------------------
# Summary

print("تم إنشاء الملفات التالية:")
for name, df in [
    ("governorates", governorates_df), ("population", population_df),
    ("hospitals", hospitals_df), ("healthcare_centers", healthcare_df),
    ("higher_education", higher_ed_df), ("education", education_df),
    ("service_facilities_by_governorate", service_fac_df),
    ("economic_establishments", econ_estab_df),
    ("economic_indicators", econ_ind_df),
]:
    print(f"  {name}.csv -> {df.shape[0]} rows x {df.shape[1]} cols")
